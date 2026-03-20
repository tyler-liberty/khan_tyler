"""
Local pacing-guide UI: runs map-pacing-guide bin scripts per job dir;
PDFs → extract text → Claude (OpenRouter) → pacing_data.json.
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.parse
import uuid
from pathlib import Path

from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent
# Load .env next to this file (not the shell cwd) so uvicorn can start from anywhere.
load_dotenv(ROOT / ".env")

from contextlib import asynccontextmanager

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
import httpx
from openai import OpenAI
from pypdf import PdfReader
from pypdf.errors import PdfReadError

OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY")
OPENROUTER_BASE_URL = os.environ.get(
    "OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1"
)
DEFAULT_MODEL = os.environ.get(
    "OPENROUTER_DEFAULT_MODEL", "anthropic/claude-sonnet-4.6"
)
MAP_PACING_SKILL_BIN = str(
    Path(
        os.environ.get(
            "MAP_PACING_SKILL_BIN",
            str(Path.home() / ".claude/skills/map-pacing-guide/bin"),
        )
    ).resolve()
)

KA_GRAPHQL_URL = os.environ.get(
    "KA_GRAPHQL_URL",
    "https://www.khanacademy.org/api/internal/graphql/GetSetOfStandards",
).rstrip("/")
KA_SITE_ORIGIN = "https://www.khanacademy.org"
KA_GRAPHQL_HASH_DEFAULT = "2804988494"


def _optional_offline_ccss_csv() -> Path | None:
    """Optional airgap path (KA_CCSS_MATH_CSV); no repo-default bundled file."""
    raw = os.environ.get("KA_CCSS_MATH_CSV", "").strip()
    if not raw:
        return None
    p = Path(raw).expanduser().resolve()
    return p if p.is_file() else None


def _normalize_standards_system(set_id: str | None) -> str:
    """Map common CCSS labels to KA's GraphQL set id CCSS.Math (pipeline + optional offline file)."""
    if set_id is None:
        return ""
    s = str(set_id).strip()
    if not s:
        return ""
    key_spaced = re.sub(r"\s+", " ", s).lower()
    key_compact = re.sub(r"\s+", "", key_spaced)
    if key_spaced in (
        "ccss.math",
        "ccss",
        "common core",
        "common core math",
        "us common core",
    ) or key_compact in ("ccssmath", "commoncore", "uscommoncore"):
        return "CCSS.Math"
    return s


def _merge_step1_form_overrides(
    pacing: dict, state_form: str, grade_form: str
) -> dict:
    """Non-empty form fields override detected state/grade (PDF or parser)."""
    out = dict(pacing)
    if state_form.strip():
        out["state"] = state_form.strip()
    if grade_form.strip():
        out["grade"] = grade_form.strip()
    return out


def _extraction_summary(
    pacing: dict, state_form: str = "", grade_form: str = ""
) -> dict:
    """What step 1 thinks jurisdiction + KA standard set are (for UI + step 2)."""
    raw_sys = pacing.get("standards_system")
    norm = _normalize_standards_system(raw_sys)
    mapped = norm if norm else raw_sys
    st = pacing.get("state")
    if st == "":
        st = None
    return {
        "state": st,
        "state_note": pacing.get("state_note"),
        "grade": pacing.get("grade"),
        "detection_confidence": pacing.get("detection_confidence"),
        "standards_system_detected": raw_sys,
        "standards_system_for_mapping": mapped,
        "standards_system_note": pacing.get("standards_system_note"),
        "used_state_override": bool(state_form.strip()),
        "used_grade_override": bool(grade_form.strip()),
    }


JOBS_ROOT = ROOT / ".tmp" / "jobs"
_http_client = httpx.Client(trust_env=False)
# KA GetSetOfStandards — large payloads; longer read timeout than OpenRouter.
_http_ka = httpx.Client(
    timeout=httpx.Timeout(120.0, connect=30.0),
    trust_env=False,
)

PDF_TO_PACING_SYSTEM = """You convert extracted text from a K-12 math pacing guide into strict RFC 8259 JSON.

Output a single JSON object with this shape (no markdown, no prose outside JSON):
{
  "state": string | null,
  "state_note": string | null,
  "standards_system": string,
  "standards_system_note": string | null,
  "grade": string,
  "detection_confidence": "high" | "medium" | "low",
  "records": [
    {
      "standard_code": string,
      "quarter": string or null,
      "date_range": string or null,
      "topic": string or null
    }
  ]
}

Identification (be explicit; users need to know what will be downloaded/mapped next):
- **state:** US two-letter postal code **only when the document clearly names that jurisdiction** (cover, standards footer, state DOE branding, or state-specific codes like MA.* for Florida BEST). If the source is national Common Core / Eureka / EngageNY with **no** state named, use **null** (unknown is fine). Do not guess a state from generic "Grade 5" alone.
- **state_note:** Optional one short phrase for humans when **state is null or ambiguous**, e.g. "Eureka modules; no state on cover", "Generic CCSS labeling", "District title only—no state". Use **null** if nothing useful to add.
- **standards_system:** This must be the **Khan Academy standard set id** the downstream mapper will use (same strings as KA’s GetSetOfStandards / CSV fetch). Prefer exactly one of: **CCSS.Math**, **FL.BEST.Math**, **IN.Math**, **TEKS.Math** when that fits. If another KA set id clearly applies, use that exact string; otherwise best match + set detection_confidence to medium/low.
- **standards_system_note:** Optional one sentence tying the document to that id, e.g. "Florida BEST codes (MA.*) → FL.BEST.Math" or "National CCSS-style codes → CCSS.Math".

Rules:
- **standard_code (critical for Khan Academy mapping):** Prefer the **official CCSS ID as used on Khan Academy (CCSS.Math)**, not curriculum shorthand.
  - Include the **cluster letter**: e.g. 5.NBT.A.1 not 5.NBT.1; 5.NBT.B.5 not 5.NBT.5; 5.NF.A.1, 5.NF.B.4, 5.NF.B.6; 5.OA.A.1, 5.OA.B.3; 5.MD.A.1, 5.MD.C.5; 5.G.A.1, 5.G.B.3.
  - Sub-standards use a **dot and lowercase letter**: e.g. 5.NF.B.4.a and 5.NF.B.4.b (not 5.NF.4a). 4.NF.B.3.c and 4.NF.B.3.d (not 4.NF.3c).
  - If the document only shows shorthand (Eureka/EngageNY style), **expand to official form** using standard CCSS notation so standard_code matches KA’s CCSS.Math CSV (e.g. map 5.NBT.1 → 5.NBT.A.1, 5.NF.4a → 5.NF.B.4.a, 4.NF.3c → 4.NF.B.3.c). When ambiguous, pick the cluster that matches the Common Core summary for that grade/domain.
- every distinct instructional standard gets one record; include module/topic/quarter/dates from the text in "topic" or "quarter" when helpful.
- If unsure about expansion, set detection_confidence to medium or low but still use the most likely official CCSS ID.
"""


def _skill_subprocess_env() -> dict[str, str]:
    """Same interpreter + PATH as this process so ka-init's `python3` sees venv packages (openpyxl, etc.)."""
    bindir = str(Path(sys.executable).resolve().parent)
    path = os.environ.get("PATH", "")
    merged = bindir + os.pathsep + path if path else bindir
    return {**os.environ, "PATH": merged}


def get_skill_bin() -> Path:
    p = Path(MAP_PACING_SKILL_BIN)
    if not (p / "ka-init").is_file():
        raise HTTPException(
            status_code=503,
            detail=(
                f"MAP_PACING_SKILL_BIN invalid or missing ka-init: "
                f"{MAP_PACING_SKILL_BIN}. Set env MAP_PACING_SKILL_BIN."
            ),
        )
    return p


def get_openai_client() -> OpenAI:
    if not OPENROUTER_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="OPENROUTER_API_KEY not configured.",
        )
    return OpenAI(
        api_key=OPENROUTER_API_KEY,
        base_url=OPENROUTER_BASE_URL.rstrip("/") + "/",
        http_client=_http_client,
    )


def _run(cmd: list[str], cwd: Path, log: list[str]) -> None:
    log.append(f"$ {' '.join(cmd)}")
    proc = subprocess.run(
        cmd,
        cwd=cwd,
        capture_output=True,
        text=True,
        env=_skill_subprocess_env(),
    )
    if proc.stdout:
        log.append(proc.stdout.rstrip())
    if proc.stderr:
        log.append(proc.stderr.rstrip())
    if proc.returncode != 0:
        raise subprocess.CalledProcessError(proc.returncode, cmd, proc.stdout, proc.stderr)


def _safe_csv_name(set_id: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9._-]+", "_", set_id).strip("_") or "standards"
    return f"{s}.csv"


def _ka_graphql_hash() -> str:
    explicit = os.environ.get("KA_GRAPHQL_HASH", "").strip()
    if explicit:
        return explicit
    cache = Path(MAP_PACING_SKILL_BIN).resolve().parent / ".ka-graphql-hash"
    if cache.is_file():
        try:
            h = cache.read_text(encoding="utf-8").strip()
            if h:
                return h
        except OSError:
            pass
    return KA_GRAPHQL_HASH_DEFAULT


def _iter_gql_mapped_rows(
    standards: list[dict], set_id: str, url_field: str
) -> list[list]:
    """Flatten GetSetOfStandards standards list to CSV rows (same as ka-standards-api)."""
    rows: list[list] = []
    url_field = (url_field or "state").strip().lower()
    for std in standards:
        mapped = std.get("mappedContent", [])
        if not mapped:
            continue
        std_code = std.get("standardId", "")
        std_desc = std.get("description", "")
        for item in mapped:
            kind = item.get("contentKind", "")
            title = item.get("title", "")
            if url_field == "default":
                url_path = item.get("defaultUrlPath", "")
            else:
                url_path = item.get("urlWithinStandardSet", "") or item.get(
                    "defaultUrlPath", ""
                )
            if url_path and not url_path.startswith("http"):
                full_url = KA_SITE_ORIGIN + url_path
            else:
                full_url = url_path or ""
            if full_url and "internal-courses" in full_url:
                continue
            rows.append([set_id, std_code, std_desc, kind, title, full_url])
    return rows


def _fetch_set_of_standards_json(set_id: str, log: list[str]) -> dict:
    graphql_hash = _ka_graphql_hash()
    variables = json.dumps({"setId": set_id, "region": "*"})
    params = urllib.parse.urlencode({"hash": graphql_hash, "variables": variables})
    url = f"{KA_GRAPHQL_URL}?{params}"
    log.append(
        f"GET KA GetSetOfStandards (hash {graphql_hash[:12]}…) for set_id={set_id!r}"
    )
    try:
        r = _http_ka.get(url, headers={"x-ka-fkey": "1"})
        r.raise_for_status()
        return r.json()
    except httpx.HTTPStatusError as e:
        log.append(f"KA GraphQL HTTP {e.response.status_code}: {e.response.text[:500]}")
        raise HTTPException(
            status_code=502,
            detail=(
                f"KA GetSetOfStandards HTTP {e.response.status_code}. "
                "Try KA_GRAPHQL_HASH from: ka-standards-api --discover-hash"
            ),
        ) from e
    except httpx.RequestError as e:
        log.append(f"KA GraphQL network error: {e}")
        raise HTTPException(
            status_code=502,
            detail=f"Could not reach Khan Academy GraphQL: {e}",
        ) from e


def _fetch_ka_mappings_temp_table(
    set_id: str,
    table_path: Path,
    log: list[str],
    url_field: str = "state",
) -> None:
    """GET GetSetOfStandards and write the tabular input ka-map-pipeline requires (--ka-csv)."""
    data = _fetch_set_of_standards_json(set_id, log)
    if "errors" in data:
        msgs = "; ".join(
            str(e.get("message", e)) for e in data.get("errors", [])
        )
        raise HTTPException(
            status_code=502,
            detail=(
                f"KA GraphQL errors: {msgs}. "
                "Update KA_GRAPHQL_HASH (run ka-standards-api --discover-hash)."
            ),
        )
    standards_data = data.get("data", {}).get("setOfStandards")
    if not standards_data:
        raise HTTPException(
            status_code=502,
            detail=(
                "KA GraphQL response missing data.setOfStandards. "
                "Hash may be stale — set KA_GRAPHQL_HASH."
            ),
        )
    top_standards = standards_data.get("standards", [])
    if not top_standards:
        top_standards = (
            standards_data if isinstance(standards_data, list) else []
        )
    rows = _iter_gql_mapped_rows(top_standards, set_id, url_field)
    header = [
        "Set ID",
        "Standard ID",
        "Standard Description",
        "Content Kind",
        "Content Title",
        "Content URL",
    ]
    table_path.parent.mkdir(parents=True, exist_ok=True)
    with table_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    n_std = sum(1 for s in top_standards if s.get("mappedContent"))
    log.append(
        f"GetSetOfStandards: {len(rows)} content rows, {n_std} standards with content "
        f"(temporary pipeline input)."
    )


def _output_xlsx_name(pacing: dict) -> str:
    grade = pacing.get("grade") or "Math"
    state = pacing.get("state") or "Unknown"
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", f"{grade}_{state}").strip("_")
    return f"{slug or 'mapping'}_KA_Mapping.xlsx"


def _format_standards_text(records: list[dict]) -> str:
    """One line per standard: code — topic — quarter (omits empty tail parts)."""
    lines: list[str] = []
    for r in records:
        code = (r.get("standard_code") or "").strip()
        if not code:
            continue
        topic = (r.get("topic") or "").strip()
        quarter = (r.get("quarter") or "").strip()
        parts = [code]
        if topic:
            parts.append(topic)
        if quarter:
            parts.append(quarter)
        lines.append(" — ".join(parts))
    return "\n".join(lines)


def _parse_standards_lines(text: str) -> list[dict]:
    """Inverse of _format_standards_text; splits on em/en dash, ASCII hyphen, or pipe."""
    records: list[dict] = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"\s*[–—|]\s*|\s+-\s+", line)
        code = parts[0].strip() if parts else ""
        if not code:
            continue
        topic = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
        quarter = parts[2].strip() if len(parts) > 2 and parts[2].strip() else None
        date_range = parts[3].strip() if len(parts) > 3 and parts[3].strip() else None
        records.append(
            {
                "standard_code": code,
                "topic": topic,
                "quarter": quarter,
                "date_range": date_range,
            }
        )
    return records


def _apply_standards_text_to_pacing(pacing_base: dict, standards_text: str) -> dict:
    records = _parse_standards_lines(standards_text)
    if not records:
        raise HTTPException(
            status_code=400,
            detail="No standards parsed from the text box. Use one line per standard: CODE — topic — quarter.",
        )
    meta = {k: v for k, v in pacing_base.items() if k != "records"}
    return {**meta, "records": records}


def _ka_init_check(job_dir: Path, upload_path: Path | None, log: list[str]) -> None:
    skill_bin = get_skill_bin()
    cmd = [str(skill_bin / "ka-init")]
    if upload_path is not None:
        cmd.append(str(upload_path))
    init_proc = subprocess.run(
        cmd,
        cwd=job_dir,
        capture_output=True,
        text=True,
        env=_skill_subprocess_env(),
    )
    init_out = (init_proc.stdout or "") + (init_proc.stderr or "")
    log.append(init_out.strip())
    if "MISSING_DEPS=" in init_out:
        raise HTTPException(
            status_code=503,
            detail=(
                "Skill dependencies missing. Install openpyxl, playwright-cli, etc. "
                "See map-pacing-guide references/setup.md — "
                f"ka-init output:\n{init_out}"
            ),
        )


def _run_mapping_stages(job_dir: Path, pacing: dict, log: list[str]) -> tuple[str, dict]:
    """Write pacing JSON, materialize KA mapping input for ka-map-pipeline, run pipeline, drop temp input."""
    skill_bin = get_skill_bin()
    py = sys.executable
    (job_dir / ".tmp").mkdir(parents=True, exist_ok=True)
    (job_dir / "output").mkdir(parents=True, exist_ok=True)

    pipeline_info: dict = {
        "standards_set_id": None,
        "ka_csv_source": None,
        "ka_csv_detail": None,
        "output_spreadsheet": None,
    }

    prev_sys = pacing.get("standards_system")
    set_id = _normalize_standards_system(prev_sys)
    # KA + ka-map-pipeline use exactly "CCSS.Math"; accept any case (e.g. CCSS.MATH).
    if (set_id or "").casefold() == "ccss.math":
        set_id = "CCSS.Math"
    if set_id != prev_sys:
        pacing = {**pacing, "standards_system": set_id}
        log.append(
            f"Standards system for mapping: {prev_sys!r} → {set_id!r}."
        )

    pacing_path = job_dir / ".tmp" / "pacing_data.json"
    pacing_path.write_text(json.dumps(pacing, indent=2), encoding="utf-8")

    if not set_id or set_id == "Unknown":
        raise HTTPException(
            status_code=422,
            detail="Could not determine standards_system in pacing data. Fix step 1 output or metadata.",
        )

    pipeline_info["standards_set_id"] = set_id

    csv_name = _safe_csv_name(set_id)
    csv_path = job_dir / ".tmp" / csv_name
    offline_ccss = _optional_offline_ccss_csv()

    try:
        if set_id == "CCSS.Math" and offline_ccss is not None:
            shutil.copy2(offline_ccss, csv_path)
            log.append(f"Using offline CCSS file (KA_CCSS_MATH_CSV): {offline_ccss}")
            pipeline_info["ka_csv_source"] = "offline_ccss_file"
            pipeline_info["ka_csv_detail"] = str(offline_ccss)
        else:
            if set_id == "CCSS.Math" and os.environ.get("KA_CCSS_MATH_CSV", "").strip():
                log.append(
                    "KA_CCSS_MATH_CSV set but file not found — using KA GraphQL for CCSS.Math."
                )
            pipeline_info["ka_csv_source"] = "ka_graphql"
            pipeline_info["ka_csv_detail"] = (
                f"GetSetOfStandards ({set_id}); temp pipeline input removed after run"
            )
            url_field = os.environ.get("KA_STANDARDS_URL_FIELD", "state").strip()
            _fetch_ka_mappings_temp_table(set_id, csv_path, log, url_field=url_field)

        out_name = _output_xlsx_name(pacing)
        out_rel = Path("output") / out_name
        _run(
            [
                py,
                str(skill_bin / "ka-map-pipeline.py"),
                "--pacing-json",
                str(pacing_path.relative_to(job_dir)),
                "--ka-csv",
                str(csv_path.relative_to(job_dir)),
                "--output",
                str(out_rel),
                "--bin-dir",
                str(skill_bin),
                "--auto-fetch",
            ],
            job_dir,
            log,
        )
        pipeline_info["output_spreadsheet"] = out_name
        pipeline_info["mapper"] = "ka-map-pipeline.py"
        return out_name, pipeline_info
    finally:
        try:
            csv_path.unlink(missing_ok=True)
        except OSError:
            pass


def _strip_markdown_json_fence(s: str) -> str:
    s = s.strip()
    if not s.startswith("```"):
        return s
    lines = s.split("\n")
    if lines and lines[0].strip().startswith("```"):
        lines = lines[1:]
    if lines and lines[-1].strip() == "```":
        lines = lines[:-1]
    return "\n".join(lines).strip()


def _extract_balanced_json_object(s: str) -> str | None:
    """First `{` through its matching `}`; handles trailing junk / prose."""
    start = s.find("{")
    if start == -1:
        return None
    depth = 0
    in_str = False
    escape = False
    quote = ""
    for i in range(start, len(s)):
        c = s[i]
        if in_str:
            if escape:
                escape = False
            elif c == "\\":
                escape = True
            elif c == quote:
                in_str = False
            continue
        if c in ('"', "'"):
            in_str = True
            quote = c
            continue
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return s[start : i + 1]
    return None


def _parse_llm_json_object(raw: str | None, log: list[str]) -> dict:
    """Normalize model output to a dict (fences, prose, empty responses)."""
    if raw is None or not str(raw).strip():
        log.append("Model returned empty content (no JSON).")
        raise HTTPException(
            status_code=502,
            detail=(
                "Model returned empty output. Try again, shorten the PDF, "
                "or use DOCX. If this persists, the router may not honor JSON mode for this model."
            ),
        )
    s = _strip_markdown_json_fence(str(raw).strip())
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        pass
    extracted = _extract_balanced_json_object(s)
    if extracted:
        try:
            return json.loads(extracted)
        except json.JSONDecodeError as e:
            log.append(f"JSON parse failed after brace extraction: {e}")
    snippet = s[:400].replace("\n", " ")
    log.append(f"Unparseable model output (first 400 chars): {snippet!r}…")
    raise HTTPException(
        status_code=502,
        detail=(
            "Model returned invalid JSON. Try again or switch model in .env "
            f"(OPENROUTER_DEFAULT_MODEL). Parse error after cleanup. Preview: {snippet[:200]}…"
        ),
    )


def pdf_text(path: Path, max_chars: int = 120_000) -> str:
    try:
        reader = PdfReader(str(path))
    except PdfReadError as e:
        raise HTTPException(status_code=400, detail=f"PDF read error: {e}") from e
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        parts.append(t)
    text = "\n\n".join(parts).strip()
    if not text:
        raise HTTPException(
            status_code=400,
            detail="No extractable text in PDF. Try exporting to DOCX or paste text into a .txt wrapper.",
        )
    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n[truncated for model context]"
    return text


_STRUCTURED_FILE_MAX_CHARS = 120_000


def structured_file_plain_text(job_dir: Path, upload_path: Path, log: list[str]) -> str:
    """DOCX/CSV/XLSX → plain text via ka-parse --emit-llm-text-to, then Claude (same as PDF)."""
    skill_bin = get_skill_bin()
    py = sys.executable
    strip_rel = Path(".tmp") / "stripped_for_llm.txt"
    _run(
        [
            py,
            str(skill_bin / "ka-parse-pacing-guide"),
            str(upload_path.resolve()),
            "--emit-llm-text-to",
            str(strip_rel),
        ],
        job_dir,
        log,
    )
    text = (job_dir / strip_rel).read_text(encoding="utf-8").strip()
    if not text:
        raise HTTPException(
            status_code=400,
            detail="No extractable text in file. Check encoding/format or try PDF.",
        )
    if len(text) > _STRUCTURED_FILE_MAX_CHARS:
        text = text[:_STRUCTURED_FILE_MAX_CHARS] + "\n\n[truncated for model context]"
    return text


def pacing_from_pdf_text(text: str, log: list[str]) -> dict:
    client = get_openai_client()
    log.append("Calling Claude via OpenRouter for PDF structuring…")
    kwargs = dict(
        model=DEFAULT_MODEL,
        messages=[
            {"role": "system", "content": PDF_TO_PACING_SYSTEM},
            {
                "role": "user",
                "content": (
                    "Pacing guide text follows (may be PDF extraction, DOCX/table text, or "
                    "tab-separated spreadsheet rows). Return only the JSON object.\n\n" + text
                ).strip(),
            },
        ],
        temperature=0.2,
    )
    # Some gateways ignore json_object; we parse defensively anyway.
    try:
        completion = client.chat.completions.create(
            **kwargs,
            response_format={"type": "json_object"},
        )
    except Exception as e:
        log.append(f"Retrying without response_format (gateway may reject json_object): {e}")
        completion = client.chat.completions.create(**kwargs)

    choice = completion.choices[0] if completion.choices else None
    if not choice:
        raise HTTPException(status_code=502, detail="Model returned no choices.")
    msg = choice.message
    refusal = getattr(msg, "refusal", None) if msg else None
    if refusal:
        log.append(f"Model refusal: {refusal}")
        raise HTTPException(status_code=502, detail=f"Model refused: {refusal}")
    raw = msg.content if msg else None
    return _parse_llm_json_object(raw, log)


@asynccontextmanager
async def lifespan(app: FastAPI):
    JOBS_ROOT.mkdir(parents=True, exist_ok=True)
    if not OPENROUTER_API_KEY:
        print("WARNING: OPENROUTER_API_KEY not set (PDF + chat will fail).")
    yield
    _http_ka.close()


app = FastAPI(title="Pacing guide web", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]
    model: str = Field(default=DEFAULT_MODEL)
    system_prompt: str | None = None
    json_mode: bool = False
    stream: bool = True


def _format_messages(messages: list[dict], system_prompt: str | None) -> list[dict]:
    formatted = [{"role": m["role"], "content": m["content"]} for m in messages]
    if system_prompt and system_prompt.strip():
        formatted.insert(0, {"role": "system", "content": system_prompt.strip()})
    return formatted


def stream_chat(
    messages: list[dict],
    model: str,
    system_prompt: str | None = None,
    json_mode: bool = False,
):
    client = get_openai_client()
    formatted = _format_messages(messages, system_prompt)
    kwargs = {
        "model": model,
        "messages": formatted,
        "stream": True,
        "temperature": 0.7,
    }
    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}
    stream = client.chat.completions.create(**kwargs)
    first = True
    for chunk in stream:
        if first:
            mu = getattr(chunk, "model", None) or model
            yield json.dumps({"model": mu, "type": "metadata"}) + "\n"
            first = False
        if chunk.choices and chunk.choices[0].delta.content:
            yield chunk.choices[0].delta.content


@app.post("/chat")
async def chat(request: ChatRequest):
    messages = [{"role": m.role, "content": m.content} for m in request.messages]
    if not messages:
        raise HTTPException(status_code=400, detail="messages cannot be empty")
    get_openai_client()
    if request.stream:
        return StreamingResponse(
            stream_chat(
                messages,
                request.model,
                request.system_prompt,
                request.json_mode,
            ),
            media_type="text/plain; charset=utf-8",
        )
    client = get_openai_client()
    kw = {
        "model": request.model,
        "messages": _format_messages(messages, request.system_prompt),
        "stream": False,
        "temperature": 0.7,
    }
    if request.json_mode:
        kw["response_format"] = {"type": "json_object"}
    completion = client.chat.completions.create(**kw)
    content = completion.choices[0].message.content or ""
    mu = getattr(completion, "model", None)
    return {"content": content, "model": mu}


@app.get("/health")
async def health():
    skill_ok = Path(MAP_PACING_SKILL_BIN) / "ka-init"
    env_path = ROOT / ".env"
    off = _optional_offline_ccss_csv()
    raw_ccss = os.environ.get("KA_CCSS_MATH_CSV", "").strip()
    return {
        "status": "ok",
        "openrouter_configured": bool(OPENROUTER_API_KEY),
        "skill_bin": MAP_PACING_SKILL_BIN,
        "skill_bin_ready": skill_ok.is_file(),
        "env_file": str(env_path),
        "env_file_present": env_path.is_file(),
        "ka_offline_ccss_path": raw_ccss or None,
        "ka_offline_ccss_usable": bool(off),
        "ka_graphql_hash_configured": bool(os.environ.get("KA_GRAPHQL_HASH", "").strip())
        or (
            Path(MAP_PACING_SKILL_BIN).resolve().parent / ".ka-graphql-hash"
        ).is_file(),
    }


@app.post("/api/step1/extract")
async def step1_extract(
    file: UploadFile = File(...),
    state: str = Form(""),
    grade: str = Form(""),
):
    """Upload → text → Claude JSON (PDF via pypdf; DOCX/CSV/XLSX via ka-parse plain text). No KA map yet."""
    job_id = str(uuid.uuid4())
    job_dir = JOBS_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=False)
    log: list[str] = []

    suffix = Path(file.filename or "upload").suffix.lower()
    if suffix not in {".docx", ".csv", ".xlsx", ".pdf"}:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Use .docx, .csv, .xlsx, or .pdf",
        )

    upload_path = job_dir / f"upload{suffix}"
    upload_path.write_bytes(await file.read())
    (job_dir / ".tmp").mkdir(exist_ok=True)
    (job_dir / "output").mkdir(exist_ok=True)

    skill_bin = get_skill_bin()
    py = sys.executable
    extracted_preview: str | None = None
    used_claude = False

    try:
        _ka_init_check(job_dir, upload_path, log)
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=str(e)) from e

    pacing_path = job_dir / ".tmp" / "pacing_data.json"
    try:
        if suffix == ".pdf":
            text = pdf_text(upload_path)
            extracted_preview = text[:8000] + ("…" if len(text) > 8000 else "")
            pacing = pacing_from_pdf_text(text, log)
            used_claude = True
            if not pacing.get("records"):
                raise HTTPException(
                    status_code=502,
                    detail="Model returned no standards records. Try another export or adjust the PDF.",
                )
        else:
            log.append(
                "Stripped plain text via ka-parse-pacing-guide --emit-llm-text-to → Claude (same JSON schema as PDF)."
            )
            text = structured_file_plain_text(job_dir, upload_path, log)
            extracted_preview = text[:8000] + ("…" if len(text) > 8000 else "")
            pacing = pacing_from_pdf_text(text, log)
            used_claude = True
            if not pacing.get("records"):
                raise HTTPException(
                    status_code=502,
                    detail="Model returned no standards records. Try a PDF or simplify the spreadsheet layout.",
                )

        if not pacing.get("records"):
            raise HTTPException(status_code=422, detail="No standards records found.")

        pacing = _merge_step1_form_overrides(pacing, state, grade)
        standards_text = _format_standards_text(pacing["records"])
        extraction_summary = _extraction_summary(pacing, state, grade)
        shutil.rmtree(job_dir, ignore_errors=True)
        return {
            "pacing_json": pacing,
            "extraction_summary": extraction_summary,
            "standards_text": standards_text,
            "extracted_text_preview": extracted_preview,
            "used_claude": used_claude,
            "log": log,
        }
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except subprocess.CalledProcessError as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(
            status_code=500,
            detail={"message": "Extract step failed", "log": log, "returncode": e.returncode},
        ) from e
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=str(e)) from e


class Step2MapRequest(BaseModel):
    pacing_base: dict = Field(
        ...,
        description="Full pacing object from step 1 (metadata preserved when merging lines).",
    )
    standards_text: str = Field(
        ...,
        description="Lines from step 2 box: CODE — topic — quarter per line.",
    )


@app.post("/api/step2/map")
async def step2_map(body: Step2MapRequest):
    """Merge edited standards lines with step 1 metadata, then run KA CSV + pipeline."""
    pacing = _apply_standards_text_to_pacing(body.pacing_base, body.standards_text)
    job_id = str(uuid.uuid4())
    job_dir = JOBS_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=False)
    log: list[str] = []

    try:
        _ka_init_check(job_dir, None, log)
        out_name, pipeline_info = _run_mapping_stages(job_dir, pacing, log)
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except subprocess.CalledProcessError as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(
            status_code=500,
            detail={"message": "Pipeline step failed", "log": log, "returncode": e.returncode},
        ) from e
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=str(e)) from e

    final_path = job_dir / "output" / out_name
    if not final_path.is_file():
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail="Output file missing after pipeline.")

    pacing_path = job_dir / ".tmp" / "pacing_data.json"
    summary = _pipeline_summary(pacing_path, log)
    return {
        "job_id": job_id,
        "download_path": f"/api/jobs/{job_id}/download",
        "filename": out_name,
        "used_pdf_llm": False,
        "pipeline": pipeline_info,
        "log": log,
        "summary": summary,
    }


@app.post("/api/run")
async def run_pipeline(
    file: UploadFile = File(...),
    state: str = Form(""),
    grade: str = Form(""),
):
    """One-shot: extract + map (legacy). Prefer /api/step1/extract then /api/step2/map."""
    skill_bin = get_skill_bin()
    job_id = str(uuid.uuid4())
    job_dir = JOBS_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=False)
    log: list[str] = []

    suffix = Path(file.filename or "upload").suffix.lower()
    if suffix not in {".docx", ".csv", ".xlsx", ".pdf"}:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Use .docx, .csv, .xlsx, or .pdf",
        )

    upload_path = job_dir / f"upload{suffix}"
    upload_path.write_bytes(await file.read())
    (job_dir / ".tmp").mkdir(exist_ok=True)
    (job_dir / "output").mkdir(exist_ok=True)

    py = sys.executable
    try:
        _ka_init_check(job_dir, upload_path, log)
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=str(e)) from e

    pacing_path = job_dir / ".tmp" / "pacing_data.json"
    use_pdf_llm = suffix == ".pdf"

    try:
        if use_pdf_llm:
            text = pdf_text(upload_path)
            pacing = pacing_from_pdf_text(text, log)
        else:
            log.append(
                "Plain text via ka-parse --emit-llm-text-to → Claude (same extract path as step1)."
            )
            text = structured_file_plain_text(job_dir, upload_path, log)
            pacing = pacing_from_pdf_text(text, log)
        if not pacing.get("records"):
            raise HTTPException(
                status_code=502,
                detail="Model returned no standards records. Try another file format or content.",
            )
        pacing_path.write_text(json.dumps(pacing, indent=2), encoding="utf-8")
        out_name, pipeline_info = _run_mapping_stages(job_dir, pacing, log)
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except subprocess.CalledProcessError as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(
            status_code=500,
            detail={"message": "Pipeline step failed", "log": log, "returncode": e.returncode},
        ) from e
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=str(e)) from e

    final_path = job_dir / "output" / out_name
    if not final_path.is_file():
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail="Output file missing after pipeline.")

    summary = _pipeline_summary(pacing_path, log)
    return {
        "job_id": job_id,
        "download_path": f"/api/jobs/{job_id}/download",
        "filename": out_name,
        "used_pdf_llm": use_pdf_llm,
        "used_claude_extract": True,
        "pipeline": pipeline_info,
        "log": log,
        "summary": summary,
    }


def _job_output_xlsx(job_dir: Path) -> Path:
    output_dir = job_dir / "output"
    files = list(output_dir.glob("*.xlsx"))
    if not files:
        raise HTTPException(status_code=404, detail="No spreadsheet output for this job")
    return max(files, key=lambda p: p.stat().st_mtime)


def _read_mapping_xlsx_preview(xlsx_path: Path, limit: int) -> list[dict]:
    """Rows from generate_xlsx: col B = standard, D = unit focus, E = exercise (+ hyperlink)."""
    from openpyxl import load_workbook

    # read_only=False so column E hyperlinks resolve to exercise_url for the UI.
    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=False)
    try:
        ws = wb.active
        last_std = ""
        last_unit = ""
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=False):
            if len(out) >= limit:
                break
            cells = list(row)
            b = cells[1].value if len(cells) > 1 else None
            d = cells[3].value if len(cells) > 3 else None
            e_cell = cells[4] if len(cells) > 4 else None
            e = e_cell.value if e_cell is not None else None
            if b is not None and str(b).strip():
                last_std = str(b).strip()
            if d is not None and str(d).strip():
                last_unit = str(d).strip()
            if e is None or not str(e).strip():
                continue
            if not last_std:
                continue
            url = None
            if e_cell is not None and e_cell.hyperlink and e_cell.hyperlink.target:
                url = e_cell.hyperlink.target
            out.append(
                {
                    "standard": last_std,
                    "unit_focus": last_unit,
                    "exercise": str(e).strip(),
                    "exercise_url": url,
                }
            )
        return out
    finally:
        wb.close()


@app.get("/api/jobs/{job_id}/preview")
async def job_preview(job_id: str, limit: int = Query(default=2000, ge=1, le=5000)):
    """Parse mapping xlsx into standard / KA unit focus / KA exercise rows for the UI."""
    job_dir = JOBS_ROOT / job_id
    if not job_dir.is_dir():
        raise HTTPException(status_code=404, detail="Job not found")
    xlsx = _job_output_xlsx(job_dir)
    rows = _read_mapping_xlsx_preview(xlsx, limit)
    return {
        "job_id": job_id,
        "source": xlsx.name,
        "row_count": len(rows),
        "truncated": len(rows) >= limit,
        "rows": rows,
    }


def _pipeline_summary(pacing_path: Path, log: list[str]) -> dict:
    # Parse last line of ka-map-pipeline if present
    text = "\n".join(log)
    out: dict = {"pacing_file": str(pacing_path.name)}
    m = re.search(r"(\d+)\s+standards", text)
    if m:
        out["standards_mentioned"] = int(m.group(1))
    return out


@app.get("/api/jobs/{job_id}/download")
async def download(job_id: str):
    job_dir = JOBS_ROOT / job_id
    if not job_dir.is_dir():
        raise HTTPException(status_code=404, detail="Job not found")
    output_dir = job_dir / "output"
    files = list(output_dir.glob("*.xlsx"))
    if not files:
        raise HTTPException(status_code=404, detail="No spreadsheet output")
    f = max(files, key=lambda p: p.stat().st_mtime)
    return FileResponse(
        path=f,
        filename=f.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


app.mount("/", StaticFiles(directory="static", html=True), name="static")
